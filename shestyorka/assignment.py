from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Side, Font, Border, Alignment

col_widths = {'A': 6.57, 'B': 64.43, 'C': 26.14, 'D': 31}
row_heights = {
    1: 27, 2: 10.5, 3: 20.25, 4: 15, 5: 15.75, 6: 15, 7: 15, 8: 18.75, 9: 24.75, 10: 24.75,
    11: 24.75, 12: 24.75, 13: 24.75, 14: 24.75, 15: 24.75, 16: 24.75, 17: 66, 18: 15, 19: 15.75, 20: 15,
    21: 15.75, 22: 15.75, 23: 15, 24: 15.75, 25: 15.75, 26: 15.75, 27: 15.75
}
table_cell_range = 'A9:D16'
merged_cells = [
    'A3:D3', 'B5:C5', 'A6:D6', 'A8:D8', 'B9:C9', 'B10:C10', 'B11:C11',
    'B12:C12', 'B13:C13', 'B14:C14', 'B15:C15', 'B16:C16', 'A17:D17',
    'B25:D25', 'B26:D26'
]
cells_with_font = {
    'A1': Font(name='Times New Roman', size=12),
    'B1': Font(name='Times New Roman', size=12),
    'A2': Font(name='Times New Roman', size=12),
    'A3': Font(name='Times New Roman', size=16, bold=True),
    'B5': Font(name='Times New Roman', size=12),
    'A6': Font(name='Times New Roman', size=10),
    'A8': Font(name='Times New Roman', size=14, bold=True),
    'A17': Font(name='Times New Roman', size=12),
    'B19': Font(name='Times New Roman', size=12, bold=True),
    'C19': Font(name='Times New Roman', size=12),
    'B22': Font(name='Times New Roman', size=12, bold=True),
    'C22': Font(name='Times New Roman', size=12),
    'B25': Font(name='Times New Roman', size=12),
    'B26': Font(name='Times New Roman', size=12)
}
cell_ranges_with_font = {
    'A9:D9': Font(name='Times New Roman', size=12, bold=True),
    'A10:D16': Font(name='Times New Roman', size=14),
    'B20:C20': Font(name='Times New Roman', size=10),
    'B23:C23': Font(name='Times New Roman', size=10)
}

cells_alignments = {
    'B19': Alignment(horizontal='left', vertical='bottom'),
    'C19': Alignment(horizontal='center', vertical='bottom'),
    'B22': Alignment(horizontal='left', vertical='bottom'),
    'C22': Alignment(horizontal='center', vertical='bottom'),
}

cell_ranges_alignments = {
    'A1:B1': Alignment(horizontal='left', vertical='bottom'),
    'A3:D3': Alignment(horizontal='center', vertical='center'),
    'B5:C5': Alignment(horizontal='left', vertical='bottom'),
    'A6:D6': Alignment(horizontal='center', vertical='bottom'),
    'A8:D8': Alignment(horizontal='left', vertical='center'),
    'A9:D9': Alignment(horizontal='center', vertical='center'),
    'A10:A16': Alignment(horizontal='center', vertical='bottom'),
    'B10:D16': Alignment(horizontal='left', vertical='center'),
    'A17:D17': Alignment(horizontal='left', vertical='bottom', wrap_text=True),
    'B20:C20': Alignment(horizontal='center', vertical='bottom'),
    'B23:C23': Alignment(horizontal='center', vertical='bottom'),
    'B25:D26': Alignment(horizontal='left', vertical='bottom')
}

cell_values = {
    'A1': 'Дата :',
    'A3': 'Наряд - задание',
    'B5': 'Выдан инспектору по надзору за освещением:                 ',
    'A6': '  (Ф.И.О.)',
    'A8': 'Поручается визуальное обследование освещенности улиц по маршруту:',
    'A9': '№ пп',
    'B9': 'Маршрут',
    'D9': 'Примечание',
    'A17': '''Во время выполнения работы:
1. При пешем обследовании находиться в сигнальном жилете.
2. Не укоснительно соблюдать ПДД для пешеходов.
3. При движении в автомобиле пользоваться ремнем безопасности.''',
    'B19': 'Наряд выдал:_____________ По Эл. Почте______________',
    'C19': 'Клыгин Д.Н.',
    'B20': 'подпись     ',
    'C20': 'фамилия, инициалы',
    'B22': 'Наряд получил:___________ По Эл. Почте______________',
    'B23': 'подпись     ',
    'C23': 'фамилия, инициалы',
    'B25': '     К работе приступил:      21:00              Работу закончил            05:00',
    'B26': '                                             время                                                     время'
}


class Assignment:
    def __init__(self, inspector_name, route_entries):
        self.book = Workbook()
        self.inspector_name = inspector_name
        self.route_entries = route_entries  # list with route entries
        # list format:
        # [{ 'name': '#category name#', 'district': '#district name#' }]
        self.sheet = self.book.active

    def _resize_col(self, col_letter, width):
        # col width = width + 0.71
        self.sheet.column_dimensions[col_letter].width = width + 0.71

    def _resize_row(self, row_num, height):
        self.sheet.row_dimensions[row_num].height = height

    def _make_table(self):
        for cells in self.sheet[table_cell_range]:
            for _cell in cells:
                self.sheet[_cell.coordinate].border = Border(
                    left=Side(border_style='thin'), right=Side(border_style='thin'),
                    top=Side(border_style='thin'), bottom=Side(border_style='thin')
                )

    def _prepare_worksheet(self):
        for col in col_widths.keys():
            self._resize_col(col, col_widths[col])

        for row in row_heights.keys():
            self._resize_row(row, row_heights[row])

        for cells in merged_cells:
            self.sheet.merge_cells(cells)

        for cell in cells_with_font:
            self.sheet[cell].font = cells_with_font[cell]

        for cell_range in cell_ranges_with_font:
            for cells in self.sheet[cell_range]:
                for cell in cells:
                    self.sheet[cell.coordinate].font = cell_ranges_with_font[cell_range]

        for cell in cells_alignments:
            self.sheet[cell].alignment = cells_alignments[cell]

        for cell_range in cell_ranges_alignments:
            for cells in self.sheet[cell_range]:
                for cell in cells:
                    self.sheet[cell.coordinate].alignment = cell_ranges_alignments[cell_range]

        for cell in cell_values:
            self.sheet[cell] = cell_values[cell]

        self._make_table()

        self.sheet['C19'].border = Border(bottom=Side(border_style='thin'))
        self.sheet['C22'].border = Border(bottom=Side(border_style='thin'))

        self.sheet.title = 'наряд'

    def _fill_table(self):
        start_row = 10
        max_entries = 7
        counter = 0
        for entry in self.route_entries:
            if counter == max_entries:
                raise OverflowError

            self.sheet.cell(start_row + counter, 1, counter + 1)
            self.sheet.cell(start_row + counter, 2, entry['name'] + ', ' + entry['district'])
            counter += 1

    def _fill_date_and_name(self):
        self.sheet.cell(5, 2).value = self.sheet.cell(5, 2).value + self.inspector_name
        self.sheet['C22'] = self.inspector_name
        today = datetime.now()
        date = ''
        date += '0' + str(today.day) if today.day < 10 else str(today.day)
        date += '.'
        date += '0' + str(today.month) if today.month < 10 else str(today.month)
        date += '.' + str(today.year)
        self.sheet['B1'] = date

    def create(self):
        self._prepare_worksheet()
        self._fill_date_and_name()
        try:
            self._fill_table()
        except OverflowError:
            return self.book, self.sheet.cell(1, 2).value, True

        return self.book, self.sheet.cell(1, 2).value, False
